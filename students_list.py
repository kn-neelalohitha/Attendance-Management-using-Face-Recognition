import pickle
import os
from os import remove
from os import getcwd
from os.path import join
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# This global variable is used to hold the list of names of students in the
# class.
names = []
# This global variable is used to hold the list of Roll numbers of students in
# the class.
rolls = []

class StudentsList:
    """ Stores the details of all students in a physical class

        Parameters
        ----------
        class_name: str
            A string that contains the name of the physical class

        Class methods
        -------------
            make_pkl_file
            load_pkl_file
            make_xl_name
            make_pkl_name
            
    """
    def __init__(self, class_name):
        """ :param str class_name: name of the physical class of a student"""
        self.class_name = class_name

    def make_pkl_file(self):
        """ Creates a Pickle(.pkl) file containing the names and roll numbers
            of all the students in the  physical class.
        """
        pkl_file_path=Path(self.make_pkl_name())
        if pkl_file_path.exists():
            os.remove(pkl_file_path)
        wb = load_workbook(self.make_xl_name())
        ws = wb.active
        number_of_studs = ws['A1'].value
        # Get the name and roll number of all the students.
        for i in range(2, number_of_studs+2):
            names.append(ws['A'+str(i)].value)
            rolls.append(ws['B'+str(i)].value)
            
        with open(self.make_pkl_name(), 'wb') as f:
            tupl = (names, rolls)
            pickle.dump(tupl, f, protocol = pickle.HIGHEST_PROTOCOL)

    def load_pkl_file(self):
        """ Reads and returns the Pickle(.pkl) file containing data of the
            physical class.
        """
        with open(self.make_pkl_name(), 'rb') as f:
            return pickle.load(f)

    def make_xl_name(self):
        """ Returns the complete pathname of the Excel(.xlsx) file containing
            data of the  physical class.
        """
        return join(getcwd(), "student's list", self.class_name + '.xlsx')
    
    def make_pkl_name(self):
        """ Returns the complete pathname of the Pickle(.pkl) file containing
            data of the physical class.
        """
        return join(getcwd(), "student's list", self.class_name + '.pkl')


if __name__ == '__main__':
	pass
